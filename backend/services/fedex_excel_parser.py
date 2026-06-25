"""
fedex_excel_parser.py — FedEx Major Incident Excel Parser
==========================================================
Reads the FedEx standard MI Excel sheet and extracts all
incident data into a structured format the RCA app understands.

Based on the exact FedEx MI sheet structure:
  - Incident Overview section (rows 1-6)
  - Business Impact section
  - Command Center Roles (MIM Lead, attendees)
  - Key Participants with team names and emails
  - Chronology of Events table
"""

import logging
from datetime import datetime, timedelta
from typing import Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class FedExMIParser:
    """
    Parses FedEx standard Major Incident Excel sheet.
    Handles the non-tabular, merged-cell format FedEx uses.
    """

    def parse(self, file_bytes: bytes) -> dict:
        """
        Main entry point. Takes raw Excel bytes, returns structured dict.
        """
        import io
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        # Read all non-empty cells into a flat list for scanning
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])

        wb.close()
        return self._extract(rows)

    def _find(self, rows, label, col_offset=1, search_col=1):
        """
        Scan rows for a label and return the value in the adjacent column.
        col_offset: how many columns to the right of the label
        """
        for row in rows:
            for i, cell in enumerate(row):
                if label.lower() in cell.lower():
                    target = i + col_offset
                    if target < len(row) and row[target]:
                        return row[target]
        return ""

    def _find_block(self, rows, label):
        """
        Find a label and return everything on that row after it.
        """
        for row in rows:
            for i, cell in enumerate(row):
                if label.lower() in cell.lower():
                    rest = [c for c in row[i+1:] if c]
                    return " ".join(rest)
        return ""

    def _extract(self, rows) -> dict:
        result = {}

        # ── Basic incident info ────────────────────────────
        inc_raw = self._find(rows, "Incident No & Priority")
        result["incident_number"] = self._parse_inc_number(inc_raw)
        result["priority"] = self._parse_priority(inc_raw)
        result["raw_inc_line"] = inc_raw

        result["start_time"]    = self._find(rows, "Incident Start Time")
        result["reported_time"] = self._find(rows, "Incident Reported")
        result["resolved_time"] = self._find(rows, "Issue Resolved Time")
        result["duration"]      = self._calc_duration(result["start_time"], result["resolved_time"])

        # ── Description & Impact ───────────────────────────
        result["issue_description"]   = self._find(rows, "Issue Description")
        result["business_impact"]     = self._find(rows, "Business Impact")
        result["any_change"]          = self._find(rows, "Any change implemented")
        result["how_reported"]        = self._find(rows, "How was the issue reported")

        # ── Impacted services ──────────────────────────────
        result["impacted_services"]   = self._find(rows, "App Name & ID")
        result["region"]              = self._find(rows, "Region Impacted")
        result["location"]            = self._find(rows, "Location Impacted")

        # ── Recovery actions ──────────────────────────────
        recovery = []
        in_recovery = False
        for row in rows:
            for i, cell in enumerate(row):
                if "Recovery Actions" in cell and "Mitigation" in cell:
                    in_recovery = True
                    break
            if in_recovery:
                action = self._find_block(rows, "Recovery Actions")
                if action:
                    recovery.append(action)
                break

        result["recovery_actions"] = self._find_block(rows, "Recovery Actions")

        # ── Command Center Roles ──────────────────────────
        result["mim_lead"]     = self._find(rows, "MIM Lead")
        result["communication"]= self._find(rows, "Communication")
        result["chronology"]   = self._find(rows, "Chronology")
        result["engaging"]     = self._find(rows, "Engaging Teams")

        # ── Escalations ────────────────────────────────────
        result["escalation_approver"] = self._find(rows, "P3 MI with Exec")

        # ── Communications timeline ────────────────────────
        result["comms_start"]  = self._find(rows, "Start", col_offset=1)
        result["comms_update"] = self._find(rows, "Update", col_offset=1)
        result["comms_end"]    = self._find(rows, "End",   col_offset=1)

        # ── Key Participants ───────────────────────────────
        result["participants"] = self._extract_participants(rows)

        # ── Chronology of events ──────────────────────────
        result["timeline_entries"] = self._extract_timeline(rows)

        # ── Build formatted timeline string ───────────────
        result["formatted_timeline"] = self._format_timeline(result["timeline_entries"])

        # ── Affected systems list ──────────────────────────
        services_raw = result.get("impacted_services", "")
        result["affected_systems"] = [
            s.strip() for s in services_raw.split(",") if s.strip()
        ]

        # ── Build description for RCA ──────────────────────
        result["full_description"] = self._build_description(result)

        # ── Map severity ───────────────────────────────────
        p = result.get("priority", "P3")
        sev_map = {"P1": "P1", "P2": "P2", "P3": "P3", "P4": "P4"}
        result["severity"] = sev_map.get(p, "P3")

        # ── Build title ────────────────────────────────────
        result["title"] = (
            f"{result['incident_number']} ({result['priority']}) — "
            f"{result.get('issue_description', 'Major Incident')[:80]}"
        )

        return result

    def _parse_inc_number(self, raw: str) -> str:
        import re
        m = re.search(r"INC\d+", raw)
        return m.group(0) if m else raw.split("(")[0].strip()

    def _parse_priority(self, raw: str) -> str:
        import re
        m = re.search(r"P[1-4]", raw)
        return m.group(0) if m else "P3"

    def _calc_duration(self, start: str, end: str) -> str:
        """Try to calculate duration from start/end time strings."""
        try:
            fmt = "%d-%m-%Y %H:%M GMT"
            s = datetime.strptime(start.strip(), fmt)
            e = datetime.strptime(end.strip(), fmt)
            delta = e - s
            mins = int(delta.total_seconds() / 60)
            return f"{mins} minutes ({mins // 60}h {mins % 60}m)"
        except Exception:
            return "Duration not calculated"

    def _extract_participants(self, rows) -> list:
        """
        Extract the Key Participants table.
        Returns list of {team, emails} dicts.
        """
        participants = []
        in_participants = False

        for row in rows:
            row_text = " ".join(row).lower()

            if "key participants" in row_text:
                in_participants = True
                continue

            if in_participants:
                # Look for rows with team name + email pattern
                non_empty = [c for c in row if c and "@" in c or (c and len(c) > 5)]
                emails_in_row = [c for c in row if c and "@" in c]
                team_in_row   = [c for c in row if c and "@" not in c and len(c) > 3 and c not in ["Action", "Information", "Completed", "Pending"]]

                if emails_in_row and team_in_row:
                    participants.append({
                        "team": team_in_row[0] if team_in_row else "Unknown",
                        "emails": [e for email in emails_in_row for e in email.split(",")]
                    })

                # Stop if we hit chronology section
                if "chronolog" in row_text and "events" in row_text:
                    break

        return participants

    def _extract_timeline(self, rows) -> list:
        """
        Extract the Chronology of Events table.
        Returns list of {date, time, action, owner, type, status} dicts.
        """
        entries = []
        in_timeline = False
        header_passed = False

        for row in rows:
            row_text = " ".join(row).lower()

            if "chronolog" in row_text and "event" in row_text:
                in_timeline = True
                continue

            if in_timeline:
                # Skip header row
                if not header_passed and ("gmt" in row_text or "timeline" in row_text):
                    header_passed = True
                    continue

                # Parse data rows — date in col 1, time in col 2, action in col 3
                non_empty = [c for c in row if c]
                if len(non_empty) < 2:
                    continue

                # Date is first non-empty, time is second
                date_val = row[1] if len(row) > 1 else ""
                time_val = row[2] if len(row) > 2 else ""
                action   = row[3] if len(row) > 3 else ""
                owner    = row[13] if len(row) > 13 else ""
                atype    = row[14] if len(row) > 14 else ""
                status   = row[15] if len(row) > 15 else ""

                # Convert Excel time fraction to HH:MM
                time_str = self._excel_time_to_str(time_val)

                if action and len(action) > 5:
                    entries.append({
                        "date": date_val,
                        "time": time_str,
                        "action": action,
                        "owner": owner,
                        "type": atype,
                        "status": status
                    })

        return entries

    def _excel_time_to_str(self, val) -> str:
        """Convert Excel fractional time (0.75 = 18:00) to HH:MM GMT."""
        try:
            f = float(val)
            total_minutes = round(f * 24 * 60)
            h = total_minutes // 60
            m = total_minutes % 60
            return f"{h:02d}:{m:02d} GMT"
        except Exception:
            return str(val)

    def _format_timeline(self, entries: list) -> str:
        """Format timeline entries into a readable string for RCA."""
        if not entries:
            return "No timeline data available."
        lines = []
        for e in entries:
            time_str = e.get("time", "")
            action   = e.get("action", "")
            owner    = e.get("owner", "")
            if action:
                line = f"{time_str} — {action}"
                if owner:
                    line += f" [{owner}]"
                lines.append(line)
        return "\n".join(lines)

    def _build_description(self, data: dict) -> str:
        """Build a rich description string for the RCA engine."""
        parts = []
        if data.get("issue_description"):
            parts.append(f"Issue: {data['issue_description']}")
        if data.get("business_impact"):
            parts.append(f"Business Impact: {data['business_impact']}")
        if data.get("impacted_services"):
            parts.append(f"Impacted Services: {data['impacted_services']}")
        if data.get("region"):
            parts.append(f"Region: {data['region']}")
        if data.get("location"):
            parts.append(f"Location: {data['location']}")
        if data.get("any_change"):
            parts.append(f"Change Implemented: {data['any_change']}")
        if data.get("how_reported"):
            parts.append(f"How Reported: {data['how_reported']}")
        if data.get("recovery_actions"):
            parts.append(f"Recovery Actions: {data['recovery_actions']}")
        if data.get("duration"):
            parts.append(f"Duration: {data['duration']}")
        return "\n\n".join(parts)

    def get_all_emails(self, parsed: dict) -> list:
        """Flatten all participant emails into a single list."""
        emails = set()
        for p in parsed.get("participants", []):
            for email in p.get("emails", []):
                email = email.strip()
                if "@" in email:
                    emails.add(email)
        return sorted(list(emails))
